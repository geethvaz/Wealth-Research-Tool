"""
Vercel Python serverless function — builds a Core Sheet Excel file.
Reads uploaded fiscal.ai xlsx files from the database, detects company type,
maps quarterly offsets, and generates a formatted Excel workbook.

Usage: GET /api/build_core_sheet?jobId=<id>
"""

from __future__ import annotations

from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import json
import os
import base64
from io import BytesIO

import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─── Constants ────────────────────────────────────────────────────────────────

FONT_NAME = "Aptos Narrow"
FONT_SIZE = 10

COLOR_BLACK = "00000000"
COLOR_DARK_GREY = "00404040"
COLOR_WHITE = "00FFFFFF"
COLOR_LIGHT_GREY = "00F2F2F2"
COLOR_GREEN = "00375623"

FMT_NUMBER = "_ * #,##0.0_ ;_ * (#,##0.0)_ ;_ * \"-\"??_ ;_ @_ "
FMT_PCT = "0.0%"
FMT_RATIO = '0.0"x"'

# ─── Database Fetcher ─────────────────────────────────────────────────────────

def _get_connection():
    """Get a database connection with proper error handling."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        raise ValueError("DATABASE_URL environment variable is not set")
    return psycopg2.connect(db_url)


def fetch_sheets(job_id: int) -> dict:
    """
    Load all uploaded xlsx files for a job from the database.
    Returns a dict mapping sheet key ('IS','CF','BS','RAT','SEG') to openpyxl Worksheet.
    """
    conn = _get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT file_type, file_data
            FROM uploaded_files
            WHERE job_id = %s AND is_screenshot = false
            """,
            (job_id,),
        )
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    type_map = {
        "income_statement": "IS",
        "cash_flow": "CF",
        "balance_sheet": "BS",
        "ratios": "RAT",
        "segments": "SEG",
    }

    sheets: dict = {}
    for file_type, file_data in rows:
        key = type_map.get(file_type)
        if key and file_data and key not in sheets:
            try:
                file_bytes = base64.b64decode(file_data)
                wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)
                sheets[key] = wb.worksheets[0]
            except Exception:
                continue  # skip corrupted files

    return sheets


def fetch_company_info(job_id: int) -> tuple:
    """Return (ticker, company_name, company_id) for the given job."""
    conn = _get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT c.ticker, c.name, c.id
            FROM build_jobs bj
            JOIN companies c ON c.id = bj.company_id
            WHERE bj.id = %s
            """,
            (job_id,),
        )
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if row:
        return row[0], row[1], row[2]
    return "UNKNOWN", "Unknown Company", None


def fetch_bull_bear(company_id: int) -> dict | None:
    """Return bull_bear JSON from core_sheets if it exists."""
    if company_id is None:
        return None
    conn = _get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT bull_bear FROM core_sheets WHERE company_id = %s LIMIT 1",
            (company_id,),
        )
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if row and row[0]:
        data = row[0]
        return data if isinstance(data, dict) else json.loads(data)
    return None


def extract_and_store_metrics(
    sheets: dict,
    offsets: dict,
    quarters: list,
    ttm_cols: dict,
    company_id: int,
    company_type: str,
):
    """Extract key financial metrics from source sheets and store in core_sheets."""
    if company_id is None:
        return

    def _val(ws, row, col):
        """Get numeric value from a cell, return None if not a number."""
        if ws is None or row is None:
            return None
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def _series(ws, label, offset, num_quarters=12):
        """Extract a series of values for a metric across quarters."""
        if ws is None:
            return {}
        row = find_row(ws, label)
        if row is None:
            return {}
        result = {}
        for q_idx in range(min(num_quarters, len(quarters))):
            src_col = (5 + q_idx) + offset
            val = _val(ws, row, src_col)
            if val is not None:
                result[quarters[q_idx] if q_idx < len(quarters) else f"Q{q_idx}"] = val
        return result

    is_ws = sheets.get("IS")
    cf_ws = sheets.get("CF")
    bs_ws = sheets.get("BS")
    rat_ws = sheets.get("RAT")
    seg_ws = sheets.get("SEG")

    is_off = offsets.get("IS", 0)
    cf_off = offsets.get("CF", 0)
    bs_off = offsets.get("BS", 0)
    rat_off = offsets.get("RAT", 0)
    seg_off = offsets.get("SEG", 0)

    if company_type == "banking":
        income_data = {
            "net_interest_income": _series(is_ws, "Net Interest Income", is_off),
            "noninterest_revenue": _series(is_ws, "Noninterest Revenue", is_off),
            "total_revenue": _series(is_ws, "Total Revenue|Total Net Revenue|Net Revenue", is_off),
            "provision_credit_losses": _series(is_ws, "Provision for Credit Losses", is_off),
            "noninterest_expense": _series(is_ws, "Noninterest Expense|Total Noninterest", is_off),
            "net_income": _series(is_ws, "Net Income", is_off),
            "efficiency_ratio": _series(is_ws, "Efficiency Ratio", is_off),
            "diluted_eps": _series(is_ws, "Diluted EPS", is_off),
        }
    else:
        income_data = {
            "revenue": _series(is_ws, "Revenue", is_off),
            "gross_profit": _series(is_ws, "Gross Profit", is_off),
            "gross_margin": _series(is_ws, "Gross Margin", is_off),
            "operating_income": _series(is_ws, "Operating Income", is_off),
            "operating_margin": _series(is_ws, "Operating Margin", is_off),
            "ebitda": _series(is_ws, "EBITDA", is_off),
            "net_income": _series(is_ws, "Net Income", is_off),
        }

    cash_flow_data = {
        "ocf": _series(cf_ws, "Operating", cf_off),
        "fcf": _series(cf_ws, "Free Cash Flow", cf_off),
        "capex": _series(cf_ws, "Capital Expenditure", cf_off),
        "sbc": _series(cf_ws, "Stock-Based", cf_off),
    }

    if company_type == "banking":
        balance_sheet_data = {
            "total_assets": _series(bs_ws, "Total Assets", bs_off),
            "total_loans": _series(bs_ws, "Total Loans|Net Loans", bs_off),
            "total_deposits": _series(bs_ws, "Total Deposits|Deposits", bs_off),
            "equity": _series(bs_ws, "Equity", bs_off),
            "tangible_book_value": _series(bs_ws, "Tangible Book Value", bs_off),
        }
        valuation_data = {
            "pe": _series(rat_ws, "P/E", rat_off),
            "pb": _series(rat_ws, "P/B", rat_off),
            "ptbv": _series(rat_ws, "P/TBV|Price/Tangible", rat_off),
            "roe": _series(rat_ws, "Return on Equity", rat_off),
            "roa": _series(rat_ws, "Return on Assets", rat_off),
        }
    else:
        balance_sheet_data = {
            "cash": _series(bs_ws, "Cash", bs_off),
            "total_assets": _series(bs_ws, "Total Assets", bs_off),
            "total_debt": _series(bs_ws, "Total Debt", bs_off),
            "equity": _series(bs_ws, "Equity", bs_off),
        }
        valuation_data = {
            "pe": _series(rat_ws, "P/E", rat_off),
            "ev_revenue": _series(rat_ws, "EV/Revenue", rat_off),
            "ev_ebitda": _series(rat_ws, "EV/EBITDA", rat_off),
            "roe": _series(rat_ws, "Return on Equity", rat_off),
            "roic": _series(rat_ws, "Return on Invested Capital", rat_off),
        }

    # Extract segment labels (first 20 rows of col A from SEG sheet)
    segments_data = {}
    if seg_ws:
        for r in range(1, min(seg_ws.max_row + 1, 40)):
            label = str(seg_ws.cell(row=r, column=1).value or "").strip()
            if label and label.lower() not in ("", "none"):
                vals = _series(seg_ws, label, seg_off)
                if vals:
                    segments_data[label] = vals

    conn = _get_connection()
    try:
        cur = conn.cursor()
        # Ensure unique index exists (idempotent)
        cur.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS core_sheets_company_id_unique ON core_sheets(company_id)"
        )
        cur.execute(
            """
            INSERT INTO core_sheets (company_id, quarters, income_statement, cash_flow,
                                     balance_sheet, valuation, segments, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            ON CONFLICT (company_id) DO UPDATE SET
                quarters = EXCLUDED.quarters,
                income_statement = EXCLUDED.income_statement,
                cash_flow = EXCLUDED.cash_flow,
                balance_sheet = EXCLUDED.balance_sheet,
                valuation = EXCLUDED.valuation,
                segments = EXCLUDED.segments,
                updated_at = NOW()
            """,
            (
                company_id,
                json.dumps(quarters),
                json.dumps(income_data),
                json.dumps(cash_flow_data),
                json.dumps(balance_sheet_data),
                json.dumps(valuation_data),
                json.dumps(segments_data),
            ),
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()


# ─── Company Type Detection ───────────────────────────────────────────────────

def detect_company_type(sheets: dict) -> str:
    is_ws = sheets.get("IS")
    seg_ws = sheets.get("SEG")

    if is_ws:
        col_a = [
            str(is_ws.cell(row=r, column=1).value or "").lower()
            for r in range(1, min(is_ws.max_row + 1, 100))
        ]
        if any("net interest income" in v for v in col_a):
            return "banking"
        if any("provision for credit losses" in v for v in col_a) and any(
            "noninterest revenue" in v for v in col_a
        ):
            return "banking"

    if seg_ws:
        col_a = [
            str(seg_ws.cell(row=r, column=1).value or "").lower()
            for r in range(1, min(seg_ws.max_row + 1, 100))
        ]
        if any("vas revenue" in v for v in col_a) or any(
            "games revenue" in v for v in col_a
        ):
            return "internet"
        if any("market intelligence" in v for v in col_a) or any(
            "ratings revenue" in v for v in col_a
        ):
            return "financials"

    return "software"


# ─── Offset Detection ─────────────────────────────────────────────────────────

def _is_date_like(val: str) -> bool:
    """Heuristic: looks like a quarter label, not TTM/NTM/blank."""
    v = val.upper().strip()
    if not v:
        return False
    skip = ("TTM", "NTM", "FORWARD", "FWD", "LTM")
    return not any(s in v for s in skip)


def detect_offsets(sheets: dict) -> tuple:
    """
    Detect the column offset for each source sheet so formulas land on the
    correct quarterly columns.

    Returns:
        offsets: dict mapping sheet key -> int offset
        quarters: list of 12 quarter label strings (oldest → newest)
        ttm_cols: dict mapping sheet key -> TTM column index (1-based), or None
        ntm_cols: dict mapping sheet key -> NTM column index (1-based), or None
    """
    is_ws = sheets.get("IS")
    if not is_ws:
        empty_offsets = {k: 0 for k in ["IS", "CF", "BS", "RAT", "SEG"]}
        empty_special = {k: None for k in ["IS", "CF", "BS", "RAT", "SEG"]}
        return empty_offsets, [], empty_special, empty_special

    # Read row 1 of IS to find date columns
    date_cols = []  # list of (col_1indexed, label_str)
    ttm_col_is = None
    ntm_col_is = None

    for col in range(1, is_ws.max_column + 1):
        val = str(is_ws.cell(row=1, column=col).value or "").strip()
        if not val:
            continue
        vu = val.upper()
        if "TTM" in vu or "LTM" in vu:
            ttm_col_is = col
        elif "NTM" in vu or "FORWARD" in vu or "FWD" in vu:
            ntm_col_is = col
        elif _is_date_like(val) and col > 1:  # skip col A (label column)
            date_cols.append((col, val))

    # Take last 12 quarterly date columns
    last_12 = date_cols[-12:] if len(date_cols) >= 12 else date_cols
    quarter_labels = [label for _, label in last_12]

    # IS offset: where does first of last-12 quarters land vs expected col 5
    is_start_col = last_12[0][0] if last_12 else 5
    is_offset = is_start_col - 5

    offsets = {"IS": is_offset}
    ttm_cols: dict = {"IS": ttm_col_is}
    ntm_cols: dict = {"IS": ntm_col_is}

    # Match dates in other sheets
    for key in ["CF", "BS", "RAT", "SEG"]:
        ws = sheets.get(key)
        if not ws or not last_12:
            offsets[key] = 0
            ttm_cols[key] = None
            ntm_cols[key] = None
            continue

        first_date = last_12[0][1]
        found_col = None
        ttm_c = None
        ntm_c = None

        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=1, column=col).value or "").strip()
            vu = val.upper()
            if "TTM" in vu or "LTM" in vu:
                ttm_c = col
            elif "NTM" in vu or "FORWARD" in vu or "FWD" in vu:
                ntm_c = col
            elif val == first_date and found_col is None:
                found_col = col

        offsets[key] = (found_col - 5) if found_col is not None else 0
        ttm_cols[key] = ttm_c
        ntm_cols[key] = ntm_c

    return offsets, quarter_labels, ttm_cols, ntm_cols


# ─── Row Finder ───────────────────────────────────────────────────────────────

def find_row(ws, search_str: str, col: int = 1) -> int | None:
    """Scan a column for a case-insensitive partial match. Returns 1-based row.
    Supports pipe-separated alternatives (e.g. 'Revenue|Total Revenues').
    Prefers rows with actual data in col B over formula-only parent rows."""
    needles = [n.strip().lower() for n in search_str.split("|")]
    candidates = []
    for row in range(1, ws.max_row + 1):
        val = str(ws.cell(row=row, column=col).value or "").lower()
        for needle in needles:
            if needle in val:
                candidates.append(row)
                break
    if not candidates:
        return None
    # Prefer rows that have data in col B (actual data vs parent/header rows)
    for r in candidates:
        v = ws.cell(row=r, column=2).value
        if v is not None:
            return r
    return candidates[0]


# ─── Sheet Copier ─────────────────────────────────────────────────────────────

def copy_sheet_data(src_ws, dest_wb: openpyxl.Workbook, sheet_name: str):
    """Copy all values from src_ws into a new sheet in dest_wb."""
    ws = dest_wb.create_sheet(sheet_name)
    for row in src_ws.iter_rows():
        for cell in row:
            dest_cell = ws.cell(row=cell.row, column=cell.column)
            dest_cell.value = cell.value
    return ws


# ─── Style Helpers ────────────────────────────────────────────────────────────

def _font(bold: bool = False, color: str = COLOR_BLACK) -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold, color=color)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _align(horizontal: str = "left", indent: int = 0) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", indent=indent)


# ─── Bull Bear Sheet Builder (shared) ────────────────────────────────────────

def _build_bull_bear_sheet(wb: openpyxl.Workbook, bull_bear: dict | None):
    """Add a 'Bull Bear & Tailwinds' sheet to the workbook."""
    bb = wb.create_sheet("Bull Bear & Tailwinds")
    bb.column_dimensions["A"].width = 80

    bb_section_map = [
        ("bull_case", "Bull Case"),
        ("bear_case", "Bear Case"),
        ("tailwinds", "Key Tailwinds"),
        ("headwinds", "Key Risks"),
        ("watchlist_metrics", "Watchlist Metrics"),
    ]

    bb_row = 1
    for json_key, section_name in bb_section_map:
        bb.merge_cells(f"A{bb_row}:H{bb_row}")
        hdr = bb[f"A{bb_row}"]
        hdr.value = section_name
        hdr.font = _font(bold=True, color=COLOR_WHITE)
        hdr.fill = _fill(COLOR_BLACK)
        hdr.alignment = _align("left", indent=1)
        bb.row_dimensions[bb_row].height = 16
        bb_row += 1

        items = (bull_bear or {}).get(json_key, [])
        if items:
            for i, item in enumerate(items):
                cell = bb.cell(row=bb_row, column=1, value=f"  {i+1}. {item}")
                cell.font = _font()
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                row_fill = _fill(COLOR_WHITE if i % 2 == 0 else COLOR_LIGHT_GREY)
                cell.fill = row_fill
                bb.row_dimensions[bb_row].height = 28
                bb_row += 1
        else:
            for _ in range(5):
                bb.row_dimensions[bb_row].height = 14
                bb_row += 1

        bb_row += 1  # gap between sections


# ─── Excel Builder: Software Template ────────────────────────────────────────

def build_software_template(
    sheets: dict,
    offsets: dict,
    quarters: list,
    ttm_cols: dict,
    ntm_cols: dict,
    ticker: str,
    company_name: str,
    bull_bear: dict | None = None,
) -> bytes:
    wb = openpyxl.Workbook()

    # Default sheet becomes Core Sheet
    cs = wb.active
    cs.title = "Core Sheet"

    # Copy source sheets after Core Sheet
    for key, name in [("IS", "IS"), ("CF", "CF"), ("BS", "BS"), ("RAT", "RAT"), ("SEG", "SEG")]:
        ws = sheets.get(key)
        if ws:
            copy_sheet_data(ws, wb, name)

    # ── View settings ──
    cs.freeze_panes = "E3"
    cs.sheet_view.zoomScale = 80

    # ── Column widths ──
    widths = {1: 38, 2: 4, 3: 8, 4: 11}
    for c in range(5, 17):   # E–P (12 quarters)
        widths[c] = 11
    widths[17] = 12           # Q = Fwd NTM
    for col_idx, w in widths.items():
        cs.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Row 1: Company header ──
    cs.merge_cells("A1:Q1")
    c1 = cs["A1"]
    c1.value = f"{company_name}  ({ticker})"
    c1.font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_BLACK)
    c1.alignment = _align("left")
    cs.row_dimensions[1].height = 18

    # ── Row 2: Period headers ──
    cs["C2"].value = "TTM"
    cs["C2"].font = _font(bold=True)
    cs["C2"].alignment = _align("center")

    for q_idx, q_label in enumerate(quarters):
        col = 5 + q_idx  # col E (5) through P (16)
        cell = cs.cell(row=2, column=col, value=q_label)
        cell.font = _font(bold=True)
        cell.alignment = _align("center")

    cs["Q2"].value = "Fwd NTM"
    cs["Q2"].font = _font(bold=True)
    cs["Q2"].alignment = _align("center")
    cs.row_dimensions[2].height = 15

    # ── Section / data row builders ──
    current_row = [3]  # mutable int via list

    def section_header(label: str):
        r = current_row[0]
        cs.merge_cells(f"A{r}:Q{r}")
        cell = cs[f"A{r}"]
        cell.value = label
        cell.font = _font(bold=True, color=COLOR_WHITE)
        cell.fill = _fill(COLOR_BLACK)
        cell.alignment = _align("left", indent=1)
        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def sub_header(label: str):
        r = current_row[0]
        cs.merge_cells(f"A{r}:Q{r}")
        cell = cs[f"A{r}"]
        cell.value = label
        cell.font = _font(bold=True, color=COLOR_WHITE)
        cell.fill = _fill(COLOR_DARK_GREY)
        cell.alignment = _align("left", indent=1)
        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def data_row(label: str, src_key: str, src_label: str, number_format: str = FMT_NUMBER):
        r = current_row[0]
        row_fill = _fill(COLOR_WHITE if (r - 3) % 2 == 0 else COLOR_LIGHT_GREY)

        # Col A: label
        a = cs.cell(row=r, column=1, value=label)
        a.font = _font()
        a.fill = row_fill

        # Fill empty cols B, D with row color
        for empty_col in [2, 4]:
            ec = cs.cell(row=r, column=empty_col)
            ec.fill = row_fill

        src_ws = sheets.get(src_key)
        src_row = find_row(src_ws, src_label) if src_ws else None

        def _rv(ws, row, col):
            v = ws.cell(row=row, column=col).value
            if v is None: return None
            if isinstance(v, (int, float)): return v
            if isinstance(v, str) and v.startswith("="): return None
            try: return float(v)
            except (ValueError, TypeError): return None

        offset = offsets.get(src_key, 0)
        vals = []

        for q_idx in range(12):
            col = 5 + q_idx
            src_col = (5 + q_idx) + offset
            cell = cs.cell(row=r, column=col)
            cell.fill = row_fill
            if src_row and src_ws:
                val = _rv(src_ws, src_row, src_col)
                if val is not None:
                    cell.value = val
                    cell.font = _font()
                    cell.number_format = number_format
                    cell.alignment = _align("right")
                    vals.append(val)

        c_ttm = cs.cell(row=r, column=3)
        c_ttm.fill = row_fill
        if src_row and src_ws:
            ttm_c = ttm_cols.get(src_key)
            if ttm_c:
                ttm_val = _rv(src_ws, src_row, ttm_c)
                if ttm_val is not None:
                    c_ttm.value = ttm_val
            elif len(vals) >= 4:
                c_ttm.value = sum(vals[-4:])
            c_ttm.font = _font()
            c_ttm.number_format = number_format
            c_ttm.alignment = _align("right")

        c_ntm = cs.cell(row=r, column=17)
        c_ntm.fill = row_fill
        if src_row and src_ws:
            ntm_c = ntm_cols.get(src_key)
            if ntm_c:
                ntm_val = _rv(src_ws, src_row, ntm_c)
                if ntm_val is not None:
                    c_ntm.value = ntm_val
            else:
                ntm_src_col = (5 + 12) + offset
                ntm_val = _rv(src_ws, src_row, ntm_src_col)
                if ntm_val is not None:
                    c_ntm.value = ntm_val
            c_ntm.font = _font()
            c_ntm.number_format = number_format
            c_ntm.alignment = _align("right")

        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def blank_row():
        cs.row_dimensions[current_row[0]].height = 6
        current_row[0] += 1

    # ════════════════════════════════════════════════════════════════════════════
    # INCOME STATEMENT
    # ════════════════════════════════════════════════════════════════════════════
    section_header("INCOME STATEMENT")
    data_row("Revenue",               "IS", "Revenue")
    data_row("Revenue Growth YoY",    "IS", "Revenue Growth",    FMT_PCT)
    data_row("Gross Profit",          "IS", "Gross Profit")
    data_row("Gross Margin",          "IS", "Gross Margin",      FMT_PCT)
    data_row("R&D",                   "IS", "Research")
    data_row("S&M / Sales & Marketing","IS", "Sales")
    data_row("G&A",                   "IS", "General")
    data_row("Operating Income",      "IS", "Operating Income")
    data_row("Operating Margin",      "IS", "Operating Margin",  FMT_PCT)
    data_row("EBITDA",                "IS", "EBITDA")
    data_row("EBITDA Margin",         "IS", "EBITDA Margin",     FMT_PCT)
    data_row("Net Income",            "IS", "Net Income")
    data_row("Net Margin",            "IS", "Net Margin",        FMT_PCT)
    data_row("Diluted EPS",           "IS", "Diluted EPS")
    data_row("Diluted Shares",        "IS", "Diluted Shares")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # CASH FLOW
    # ════════════════════════════════════════════════════════════════════════════
    section_header("CASH FLOW")
    data_row("Cash from Operations",      "CF", "Operating")
    data_row("Capital Expenditures",      "CF", "Capital Expenditure")
    data_row("Free Cash Flow",            "CF", "Free Cash Flow")
    data_row("FCF Margin",                "CF", "FCF Margin",         FMT_PCT)
    data_row("Stock-Based Compensation",  "CF", "Stock-Based")
    data_row("Acquisitions",              "CF", "Acquisition")
    data_row("Share Buybacks",            "CF", "Repurchase")
    data_row("Dividends Paid",            "CF", "Dividend")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # BALANCE SHEET
    # ════════════════════════════════════════════════════════════════════════════
    section_header("BALANCE SHEET")
    data_row("Cash & Equivalents",     "BS", "Cash")
    data_row("Total Assets",           "BS", "Total Assets")
    data_row("Total Debt",             "BS", "Total Debt")
    data_row("Net Cash / (Debt)",      "BS", "Net Cash")
    data_row("Shareholders' Equity",   "BS", "Equity")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # VALUATION
    # ════════════════════════════════════════════════════════════════════════════
    section_header("VALUATION")

    sub_header("Price Multiples")
    data_row("P/E",             "RAT", "P/E",   FMT_RATIO)
    data_row("P/FCF",           "RAT", "P/FCF", FMT_RATIO)
    data_row("P/S",             "RAT", "P/S",   FMT_RATIO)

    sub_header("EV Multiples")
    data_row("EV / Revenue",    "RAT", "EV/Revenue", FMT_RATIO)
    data_row("EV / EBITDA",     "RAT", "EV/EBITDA",  FMT_RATIO)
    data_row("EV / EBIT",       "RAT", "EV/EBIT",    FMT_RATIO)
    data_row("EV / FCF",        "RAT", "EV/FCF",     FMT_RATIO)

    sub_header("Returns")
    data_row("Return on Equity (ROE)",          "RAT", "Return on Equity",          FMT_PCT)
    data_row("Return on Assets (ROA)",          "RAT", "Return on Assets",          FMT_PCT)
    data_row("Return on Invested Capital (ROIC)","RAT","Return on Invested Capital", FMT_PCT)
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # BULL BEAR & TAILWINDS
    # ════════════════════════════════════════════════════════════════════════════
    _build_bull_bear_sheet(wb, bull_bear)

    # ── Serialize ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ─── Excel Builder: Banking Template ────────────────────────────────────────

def build_banking_template(
    sheets: dict,
    offsets: dict,
    quarters: list,
    ttm_cols: dict,
    ntm_cols: dict,
    ticker: str,
    company_name: str,
    bull_bear: dict | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    cs = wb.active
    cs.title = "Core Sheet"

    for key, name in [("IS", "IS"), ("CF", "CF"), ("BS", "BS"), ("RAT", "RAT"), ("SEG", "SEG")]:
        ws = sheets.get(key)
        if ws:
            copy_sheet_data(ws, wb, name)

    cs.freeze_panes = "E3"
    cs.sheet_view.zoomScale = 80

    widths = {1: 38, 2: 4, 3: 8, 4: 11}
    for c in range(5, 17):
        widths[c] = 11
    widths[17] = 12
    for col_idx, w in widths.items():
        cs.column_dimensions[get_column_letter(col_idx)].width = w

    cs.merge_cells("A1:Q1")
    c1 = cs["A1"]
    c1.value = f"{company_name}  ({ticker})"
    c1.font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_BLACK)
    c1.alignment = _align("left")
    cs.row_dimensions[1].height = 18

    cs["C2"].value = "TTM"
    cs["C2"].font = _font(bold=True)
    cs["C2"].alignment = _align("center")

    for q_idx, q_label in enumerate(quarters):
        col = 5 + q_idx
        cell = cs.cell(row=2, column=col, value=q_label)
        cell.font = _font(bold=True)
        cell.alignment = _align("center")

    cs["Q2"].value = "Fwd NTM"
    cs["Q2"].font = _font(bold=True)
    cs["Q2"].alignment = _align("center")
    cs.row_dimensions[2].height = 15

    current_row = [3]

    def _read_val_b(ws, row, col):
        """Read actual numeric value from a cell (not formula)."""
        if ws is None or row is None:
            return None
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def section_header(label: str):
        r = current_row[0]
        cs.merge_cells(f"A{r}:Q{r}")
        cell = cs[f"A{r}"]
        cell.value = label
        cell.font = _font(bold=True, color=COLOR_WHITE)
        cell.fill = _fill(COLOR_BLACK)
        cell.alignment = _align("left", indent=1)
        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def sub_header(label: str):
        r = current_row[0]
        cs.merge_cells(f"A{r}:Q{r}")
        cell = cs[f"A{r}"]
        cell.value = label
        cell.font = _font(bold=True, color=COLOR_WHITE)
        cell.fill = _fill(COLOR_DARK_GREY)
        cell.alignment = _align("left", indent=1)
        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def data_row(label: str, src_key: str, src_label: str, number_format: str = FMT_NUMBER):
        r = current_row[0]
        row_fill = _fill(COLOR_WHITE if (r - 3) % 2 == 0 else COLOR_LIGHT_GREY)

        a = cs.cell(row=r, column=1, value=label)
        a.font = _font()
        a.fill = row_fill

        for empty_col in [2, 4]:
            ec = cs.cell(row=r, column=empty_col)
            ec.fill = row_fill

        src_ws = sheets.get(src_key)
        src_row = find_row(src_ws, src_label) if src_ws else None

        def _rv(ws, row, col):
            v = ws.cell(row=row, column=col).value
            if v is None: return None
            if isinstance(v, (int, float)): return v
            if isinstance(v, str) and v.startswith("="): return None
            try: return float(v)
            except (ValueError, TypeError): return None

        offset = offsets.get(src_key, 0)
        vals = []

        # 12 quarters — write actual values
        for q_idx in range(12):
            col = 5 + q_idx
            src_col = (5 + q_idx) + offset
            cell = cs.cell(row=r, column=col)
            cell.fill = row_fill
            if src_row and src_ws:
                val = _rv(src_ws, src_row, src_col)
                if val is not None:
                    cell.value = val
                    cell.font = _font()
                    cell.number_format = number_format
                    cell.alignment = _align("right")
                    vals.append(val)

        # TTM
        c_ttm = cs.cell(row=r, column=3)
        c_ttm.fill = row_fill
        if src_row and src_ws:
            ttm_c = ttm_cols.get(src_key)
            if ttm_c:
                ttm_val = _rv(src_ws, src_row, ttm_c)
                if ttm_val is not None:
                    c_ttm.value = ttm_val
            elif len(vals) >= 4:
                c_ttm.value = sum(vals[-4:])
            c_ttm.font = _font()
            c_ttm.number_format = number_format
            c_ttm.alignment = _align("right")

        # NTM
        c_ntm = cs.cell(row=r, column=17)
        c_ntm.fill = row_fill
        if src_row and src_ws:
            ntm_c = ntm_cols.get(src_key)
            if ntm_c:
                ntm_val = _rv(src_ws, src_row, ntm_c)
                if ntm_val is not None:
                    c_ntm.value = ntm_val
            else:
                ntm_src_col = (5 + 12) + offset
                ntm_val = _rv(src_ws, src_row, ntm_src_col)
                if ntm_val is not None:
                    c_ntm.value = ntm_val
            c_ntm.font = _font()
            c_ntm.number_format = number_format
            c_ntm.alignment = _align("right")

        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def blank_row():
        cs.row_dimensions[current_row[0]].height = 6
        current_row[0] += 1

    # ════════════════════════════════════════════════════════════════════════════
    # INCOME STATEMENT (Banking)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("INCOME STATEMENT")
    data_row("Net Interest Income",         "IS", "Net Interest Income")
    data_row("Noninterest Revenue",         "IS", "Noninterest Revenue")
    data_row("Total Net Revenue",           "IS", "Total Revenue|Total Net Revenue|Net Revenue")
    data_row("Revenue Growth YoY",          "IS", "Revenue Growth", FMT_PCT)
    data_row("Provision for Credit Losses", "IS", "Provision for Credit Losses")
    data_row("Noninterest Expense",         "IS", "Noninterest Expense|Total Noninterest")
    data_row("Pre-Tax Income",              "IS", "Pre-Tax Income|Income Before Tax")
    data_row("Net Income",                  "IS", "Net Income")
    data_row("Net Margin",                  "IS", "Net Margin", FMT_PCT)
    data_row("Diluted EPS",                 "IS", "Diluted EPS")
    data_row("Diluted Shares",              "IS", "Diluted Shares")
    data_row("Efficiency Ratio",            "IS", "Efficiency Ratio", FMT_PCT)
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # CASH FLOW (Banking)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("CASH FLOW")
    data_row("Cash from Operations",      "CF", "Operating")
    data_row("Capital Expenditures",      "CF", "Capital Expenditure")
    data_row("Dividends Paid",            "CF", "Dividend")
    data_row("Share Buybacks",            "CF", "Repurchase")
    data_row("Acquisitions",              "CF", "Acquisition")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # BALANCE SHEET (Banking)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("BALANCE SHEET")
    data_row("Total Assets",              "BS", "Total Assets")
    data_row("Total Loans, Net",          "BS", "Total Loans|Net Loans")
    data_row("Total Deposits",            "BS", "Total Deposits|Deposits")
    data_row("Total Debt",                "BS", "Total Debt")
    data_row("Shareholders' Equity",      "BS", "Equity")
    data_row("Tangible Book Value",       "BS", "Tangible Book Value")
    data_row("Book Value per Share",      "BS", "Book Value Per Share|BVPS")
    data_row("Tangible BV per Share",     "BS", "Tangible Book Value Per Share|TBVPS")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # SEGMENTS (Banking — auto-detect from SEG sheet)
    # ════════════════════════════════════════════════════════════════════════════
    seg_ws = sheets.get("SEG")
    if seg_ws:
        section_header("SEGMENTS")
        seg_labels = []
        for r in range(1, min(seg_ws.max_row + 1, 80)):
            label = str(seg_ws.cell(row=r, column=1).value or "").strip()
            if label and label.lower() not in ("", "none"):
                seg_labels.append(label)
        for lbl in seg_labels:
            # Detect format from label name
            fmt = FMT_NUMBER
            lbl_lower = lbl.lower()
            if "margin" in lbl_lower or "ratio" in lbl_lower or "growth" in lbl_lower or "mix" in lbl_lower:
                fmt = FMT_PCT
            data_row(lbl, "SEG", lbl, fmt)
        blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # VALUATION (Banking — P/B, P/TBV instead of EV multiples)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("VALUATION")

    sub_header("Price Multiples")
    data_row("P/E",                "RAT", "P/E",   FMT_RATIO)
    data_row("P/B",                "RAT", "P/B",   FMT_RATIO)
    data_row("P/TBV",             "RAT", "P/TBV|Price/Tangible", FMT_RATIO)

    sub_header("Returns")
    data_row("Return on Equity (ROE)",    "RAT", "Return on Equity",    FMT_PCT)
    data_row("Return on Assets (ROA)",    "RAT", "Return on Assets",    FMT_PCT)
    data_row("Return on Tangible Equity", "RAT", "Return on Tangible|ROTE", FMT_PCT)
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # BULL BEAR & TAILWINDS
    # ════════════════════════════════════════════════════════════════════════════
    _build_bull_bear_sheet(wb, bull_bear)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ─── Excel Builder: Financials Template ────────────────────────────────────

def build_financials_template(
    sheets: dict,
    offsets: dict,
    quarters: list,
    ttm_cols: dict,
    ntm_cols: dict,
    ticker: str,
    company_name: str,
    bull_bear: dict | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    cs = wb.active
    cs.title = "Core Sheet"

    for key, name in [("IS", "IS"), ("CF", "CF"), ("BS", "BS"), ("RAT", "RAT"), ("SEG", "SEG")]:
        ws = sheets.get(key)
        if ws:
            copy_sheet_data(ws, wb, name)

    cs.freeze_panes = "E3"
    cs.sheet_view.zoomScale = 80

    widths = {1: 38, 2: 4, 3: 8, 4: 11}
    for c in range(5, 17):
        widths[c] = 11
    widths[17] = 12
    for col_idx, w in widths.items():
        cs.column_dimensions[get_column_letter(col_idx)].width = w

    cs.merge_cells("A1:Q1")
    c1 = cs["A1"]
    c1.value = f"{company_name}  ({ticker})"
    c1.font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_BLACK)
    c1.alignment = _align("left")
    cs.row_dimensions[1].height = 18

    cs["C2"].value = "TTM"
    cs["C2"].font = _font(bold=True)
    cs["C2"].alignment = _align("center")

    for q_idx, q_label in enumerate(quarters):
        col = 5 + q_idx
        cell = cs.cell(row=2, column=col, value=q_label)
        cell.font = _font(bold=True)
        cell.alignment = _align("center")

    cs["Q2"].value = "Fwd NTM"
    cs["Q2"].font = _font(bold=True)
    cs["Q2"].alignment = _align("center")
    cs.row_dimensions[2].height = 15

    current_row = [3]

    def section_header(label: str):
        r = current_row[0]
        cs.merge_cells(f"A{r}:Q{r}")
        cell = cs[f"A{r}"]
        cell.value = label
        cell.font = _font(bold=True, color=COLOR_WHITE)
        cell.fill = _fill(COLOR_BLACK)
        cell.alignment = _align("left", indent=1)
        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def sub_header(label: str):
        r = current_row[0]
        cs.merge_cells(f"A{r}:Q{r}")
        cell = cs[f"A{r}"]
        cell.value = label
        cell.font = _font(bold=True, color=COLOR_WHITE)
        cell.fill = _fill(COLOR_DARK_GREY)
        cell.alignment = _align("left", indent=1)
        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def data_row(label: str, src_key: str, src_label: str, number_format: str = FMT_NUMBER):
        r = current_row[0]
        row_fill = _fill(COLOR_WHITE if (r - 3) % 2 == 0 else COLOR_LIGHT_GREY)

        a = cs.cell(row=r, column=1, value=label)
        a.font = _font()
        a.fill = row_fill

        for empty_col in [2, 4]:
            ec = cs.cell(row=r, column=empty_col)
            ec.fill = row_fill

        src_ws = sheets.get(src_key)
        src_row = find_row(src_ws, src_label) if src_ws else None

        def _rv(ws, row, col):
            v = ws.cell(row=row, column=col).value
            if v is None: return None
            if isinstance(v, (int, float)): return v
            if isinstance(v, str) and v.startswith("="): return None
            try: return float(v)
            except (ValueError, TypeError): return None

        offset = offsets.get(src_key, 0)
        vals = []

        # 12 quarters — write actual values
        for q_idx in range(12):
            col = 5 + q_idx
            src_col = (5 + q_idx) + offset
            cell = cs.cell(row=r, column=col)
            cell.fill = row_fill
            if src_row and src_ws:
                val = _rv(src_ws, src_row, src_col)
                if val is not None:
                    cell.value = val
                    cell.font = _font()
                    cell.number_format = number_format
                    cell.alignment = _align("right")
                    vals.append(val)

        # TTM
        c_ttm = cs.cell(row=r, column=3)
        c_ttm.fill = row_fill
        if src_row and src_ws:
            ttm_c = ttm_cols.get(src_key)
            if ttm_c:
                ttm_val = _rv(src_ws, src_row, ttm_c)
                if ttm_val is not None:
                    c_ttm.value = ttm_val
            elif len(vals) >= 4:
                c_ttm.value = sum(vals[-4:])
            c_ttm.font = _font()
            c_ttm.number_format = number_format
            c_ttm.alignment = _align("right")

        # NTM
        c_ntm = cs.cell(row=r, column=17)
        c_ntm.fill = row_fill
        if src_row and src_ws:
            ntm_c = ntm_cols.get(src_key)
            if ntm_c:
                ntm_val = _rv(src_ws, src_row, ntm_c)
                if ntm_val is not None:
                    c_ntm.value = ntm_val
            else:
                ntm_src_col = (5 + 12) + offset
                ntm_val = _rv(src_ws, src_row, ntm_src_col)
                if ntm_val is not None:
                    c_ntm.value = ntm_val
            c_ntm.font = _font()
            c_ntm.number_format = number_format
            c_ntm.alignment = _align("right")

        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def blank_row():
        cs.row_dimensions[current_row[0]].height = 6
        current_row[0] += 1

    # ════════════════════════════════════════════════════════════════════════════
    # INCOME STATEMENT (Financials — SPGI style)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("INCOME STATEMENT")
    data_row("Revenue",               "IS", "Revenue|Total Revenue")
    data_row("Revenue Growth YoY",    "IS", "Revenue Growth",    FMT_PCT)
    data_row("Operating Expenses",    "IS", "Operating Expense|Total Operating")
    data_row("Operating Income",      "IS", "Operating Income")
    data_row("Operating Margin",      "IS", "Operating Margin",  FMT_PCT)
    data_row("EBITDA",                "IS", "EBITDA")
    data_row("EBITDA Margin",         "IS", "EBITDA Margin",     FMT_PCT)
    data_row("Interest Expense",      "IS", "Interest Expense")
    data_row("Net Income",            "IS", "Net Income")
    data_row("Net Margin",            "IS", "Net Margin",        FMT_PCT)
    data_row("Diluted EPS",           "IS", "Diluted EPS")
    data_row("Diluted Shares",        "IS", "Diluted Shares")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # CASH FLOW (Financials)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("CASH FLOW")
    data_row("Cash from Operations",      "CF", "Operating")
    data_row("Capital Expenditures",      "CF", "Capital Expenditure")
    data_row("Free Cash Flow",            "CF", "Free Cash Flow")
    data_row("FCF Margin",                "CF", "FCF Margin",         FMT_PCT)
    data_row("Stock-Based Compensation",  "CF", "Stock-Based")
    data_row("Acquisitions",              "CF", "Acquisition")
    data_row("Share Buybacks",            "CF", "Repurchase")
    data_row("Dividends Paid",            "CF", "Dividend")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # SEGMENTS (Financials — auto-detect from SEG sheet)
    # e.g. Market Intelligence, Ratings, Commodity Insights, Mobility, Indices
    # ════════════════════════════════════════════════════════════════════════════
    seg_ws = sheets.get("SEG")
    if seg_ws:
        section_header("SEGMENTS")

        # Scan SEG sheet for segment labels and group them
        seg_labels = []
        for r in range(1, min(seg_ws.max_row + 1, 120)):
            label = str(seg_ws.cell(row=r, column=1).value or "").strip()
            if label and label.lower() not in ("", "none"):
                seg_labels.append(label)

        # Group by revenue segments, then operating profit, then margins
        rev_labels = []
        op_labels = []
        margin_labels = []
        growth_labels = []
        mix_labels = []
        other_labels = []

        for lbl in seg_labels:
            ll = lbl.lower()
            if "revenue" in ll and "growth" not in ll and "mix" not in ll:
                rev_labels.append(lbl)
            elif "operating profit" in ll or "operating income" in ll:
                op_labels.append(lbl)
            elif "margin" in ll:
                margin_labels.append(lbl)
            elif "growth" in ll:
                growth_labels.append(lbl)
            elif "mix" in ll or "% of" in ll:
                mix_labels.append(lbl)
            else:
                other_labels.append(lbl)

        if rev_labels:
            sub_header("Segment Revenue")
            for lbl in rev_labels:
                data_row(lbl, "SEG", lbl)

        if op_labels:
            sub_header("Segment Operating Profit")
            for lbl in op_labels:
                data_row(lbl, "SEG", lbl)

        if margin_labels:
            sub_header("Segment Margins")
            for lbl in margin_labels:
                data_row(lbl, "SEG", lbl, FMT_PCT)

        if growth_labels:
            sub_header("Segment Growth")
            for lbl in growth_labels:
                data_row(lbl, "SEG", lbl, FMT_PCT)

        if mix_labels:
            sub_header("Revenue Mix")
            for lbl in mix_labels:
                data_row(lbl, "SEG", lbl, FMT_PCT)

        if other_labels:
            sub_header("Other KPIs")
            for lbl in other_labels:
                data_row(lbl, "SEG", lbl)

        blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # BALANCE SHEET (Financials)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("BALANCE SHEET")
    data_row("Cash & Equivalents",     "BS", "Cash")
    data_row("Total Assets",           "BS", "Total Assets")
    data_row("Goodwill & Intangibles", "BS", "Goodwill|Intangible")
    data_row("Total Debt",             "BS", "Total Debt")
    data_row("Net Cash / (Debt)",      "BS", "Net Cash")
    data_row("Shareholders' Equity",   "BS", "Equity")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # VALUATION (Financials)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("VALUATION")

    sub_header("Price Multiples")
    data_row("P/E",             "RAT", "P/E",   FMT_RATIO)
    data_row("P/FCF",           "RAT", "P/FCF", FMT_RATIO)
    data_row("P/S",             "RAT", "P/S",   FMT_RATIO)

    sub_header("EV Multiples")
    data_row("EV / Revenue",    "RAT", "EV/Revenue", FMT_RATIO)
    data_row("EV / EBITDA",     "RAT", "EV/EBITDA",  FMT_RATIO)
    data_row("EV / EBIT",       "RAT", "EV/EBIT",    FMT_RATIO)
    data_row("EV / FCF",        "RAT", "EV/FCF",     FMT_RATIO)

    sub_header("Returns")
    data_row("Return on Equity (ROE)",          "RAT", "Return on Equity",          FMT_PCT)
    data_row("Return on Assets (ROA)",          "RAT", "Return on Assets",          FMT_PCT)
    data_row("Return on Invested Capital (ROIC)","RAT","Return on Invested Capital", FMT_PCT)
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # BULL BEAR & TAILWINDS
    # ════════════════════════════════════════════════════════════════════════════
    _build_bull_bear_sheet(wb, bull_bear)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ─── Excel Builder: Internet Template ──────────────────────────────────────

def build_internet_template(
    sheets: dict,
    offsets: dict,
    quarters: list,
    ttm_cols: dict,
    ntm_cols: dict,
    ticker: str,
    company_name: str,
    bull_bear: dict | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    cs = wb.active
    cs.title = "Core Sheet"

    for key, name in [("IS", "IS"), ("CF", "CF"), ("BS", "BS"), ("RAT", "RAT"), ("SEG", "SEG")]:
        ws = sheets.get(key)
        if ws:
            copy_sheet_data(ws, wb, name)

    cs.freeze_panes = "E3"
    cs.sheet_view.zoomScale = 80

    widths = {1: 38, 2: 4, 3: 8, 4: 11}
    for c in range(5, 17):
        widths[c] = 11
    widths[17] = 12
    for col_idx, w in widths.items():
        cs.column_dimensions[get_column_letter(col_idx)].width = w

    cs.merge_cells("A1:Q1")
    c1 = cs["A1"]
    c1.value = f"{company_name}  ({ticker})"
    c1.font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_BLACK)
    c1.alignment = _align("left")
    cs.row_dimensions[1].height = 18

    cs["C2"].value = "TTM"
    cs["C2"].font = _font(bold=True)
    cs["C2"].alignment = _align("center")

    for q_idx, q_label in enumerate(quarters):
        col = 5 + q_idx
        cell = cs.cell(row=2, column=col, value=q_label)
        cell.font = _font(bold=True)
        cell.alignment = _align("center")

    cs["Q2"].value = "Fwd NTM"
    cs["Q2"].font = _font(bold=True)
    cs["Q2"].alignment = _align("center")
    cs.row_dimensions[2].height = 15

    current_row = [3]

    def _read_val_b(ws, row, col):
        """Read actual numeric value from a cell (not formula)."""
        if ws is None or row is None:
            return None
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def section_header(label: str):
        r = current_row[0]
        cs.merge_cells(f"A{r}:Q{r}")
        cell = cs[f"A{r}"]
        cell.value = label
        cell.font = _font(bold=True, color=COLOR_WHITE)
        cell.fill = _fill(COLOR_BLACK)
        cell.alignment = _align("left", indent=1)
        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def sub_header(label: str):
        r = current_row[0]
        cs.merge_cells(f"A{r}:Q{r}")
        cell = cs[f"A{r}"]
        cell.value = label
        cell.font = _font(bold=True, color=COLOR_WHITE)
        cell.fill = _fill(COLOR_DARK_GREY)
        cell.alignment = _align("left", indent=1)
        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def data_row(label: str, src_key: str, src_label: str, number_format: str = FMT_NUMBER):
        r = current_row[0]
        row_fill = _fill(COLOR_WHITE if (r - 3) % 2 == 0 else COLOR_LIGHT_GREY)

        a = cs.cell(row=r, column=1, value=label)
        a.font = _font()
        a.fill = row_fill

        for empty_col in [2, 4]:
            ec = cs.cell(row=r, column=empty_col)
            ec.fill = row_fill

        src_ws = sheets.get(src_key)
        src_row = find_row(src_ws, src_label) if src_ws else None

        def _rv(ws, row, col):
            v = ws.cell(row=row, column=col).value
            if v is None: return None
            if isinstance(v, (int, float)): return v
            if isinstance(v, str) and v.startswith("="): return None
            try: return float(v)
            except (ValueError, TypeError): return None

        offset = offsets.get(src_key, 0)
        vals = []

        # 12 quarters — write actual values
        for q_idx in range(12):
            col = 5 + q_idx
            src_col = (5 + q_idx) + offset
            cell = cs.cell(row=r, column=col)
            cell.fill = row_fill
            if src_row and src_ws:
                val = _rv(src_ws, src_row, src_col)
                if val is not None:
                    cell.value = val
                    cell.font = _font()
                    cell.number_format = number_format
                    cell.alignment = _align("right")
                    vals.append(val)

        # TTM
        c_ttm = cs.cell(row=r, column=3)
        c_ttm.fill = row_fill
        if src_row and src_ws:
            ttm_c = ttm_cols.get(src_key)
            if ttm_c:
                ttm_val = _rv(src_ws, src_row, ttm_c)
                if ttm_val is not None:
                    c_ttm.value = ttm_val
            elif len(vals) >= 4:
                c_ttm.value = sum(vals[-4:])
            c_ttm.font = _font()
            c_ttm.number_format = number_format
            c_ttm.alignment = _align("right")

        # NTM
        c_ntm = cs.cell(row=r, column=17)
        c_ntm.fill = row_fill
        if src_row and src_ws:
            ntm_c = ntm_cols.get(src_key)
            if ntm_c:
                ntm_val = _rv(src_ws, src_row, ntm_c)
                if ntm_val is not None:
                    c_ntm.value = ntm_val
            else:
                ntm_src_col = (5 + 12) + offset
                ntm_val = _rv(src_ws, src_row, ntm_src_col)
                if ntm_val is not None:
                    c_ntm.value = ntm_val
            c_ntm.font = _font()
            c_ntm.number_format = number_format
            c_ntm.alignment = _align("right")

        cs.row_dimensions[r].height = 14
        current_row[0] += 1

    def blank_row():
        cs.row_dimensions[current_row[0]].height = 6
        current_row[0] += 1

    # ════════════════════════════════════════════════════════════════════════════
    # INCOME STATEMENT (Internet — Tencent style)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("INCOME STATEMENT")
    data_row("Revenue",               "IS", "Revenue|Total Revenue|Revenues")
    data_row("Revenue Growth YoY",    "IS", "Revenue Growth",    FMT_PCT)
    data_row("Cost of Revenue",       "IS", "Cost of Revenue|Cost of Sales")
    data_row("Gross Profit",          "IS", "Gross Profit")
    data_row("Gross Margin",          "IS", "Gross Margin",      FMT_PCT)
    data_row("Selling & Marketing",   "IS", "Selling|Marketing|Sales")
    data_row("G&A",                   "IS", "General")
    data_row("R&D",                   "IS", "Research")
    data_row("Operating Income",      "IS", "Operating Income")
    data_row("Operating Margin",      "IS", "Operating Margin",  FMT_PCT)
    data_row("EBITDA",                "IS", "EBITDA")
    data_row("EBITDA Margin",         "IS", "EBITDA Margin",     FMT_PCT)
    data_row("Net Income",            "IS", "Net Income")
    data_row("Net Margin",            "IS", "Net Margin",        FMT_PCT)
    data_row("Diluted EPS",           "IS", "Diluted EPS")
    data_row("Diluted Shares",        "IS", "Diluted Shares")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # CASH FLOW (Internet)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("CASH FLOW")
    data_row("Cash from Operations",      "CF", "Operating")
    data_row("Capital Expenditures",      "CF", "Capital Expenditure")
    data_row("Free Cash Flow",            "CF", "Free Cash Flow")
    data_row("FCF Margin",                "CF", "FCF Margin",         FMT_PCT)
    data_row("Stock-Based Compensation",  "CF", "Stock-Based")
    data_row("Acquisitions",              "CF", "Acquisition")
    data_row("Share Buybacks",            "CF", "Repurchase")
    data_row("Dividends Paid",            "CF", "Dividend")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # SEGMENTS (Internet — auto-detect from SEG sheet)
    # e.g. VAS Revenue, Games Revenue, Online Advertising, FinTech & Business Services
    # ════════════════════════════════════════════════════════════════════════════
    seg_ws = sheets.get("SEG")
    if seg_ws:
        section_header("SEGMENTS & KPIs")

        seg_labels = []
        for r in range(1, min(seg_ws.max_row + 1, 120)):
            label = str(seg_ws.cell(row=r, column=1).value or "").strip()
            if label and label.lower() not in ("", "none"):
                seg_labels.append(label)

        # Group by revenue, user metrics (MAU/DAU), growth, mix, other
        rev_labels = []
        user_labels = []
        growth_labels = []
        mix_labels = []
        other_labels = []

        for lbl in seg_labels:
            ll = lbl.lower()
            if "revenue" in ll and "growth" not in ll and "mix" not in ll:
                rev_labels.append(lbl)
            elif "mau" in ll or "dau" in ll or "user" in ll or "subscriber" in ll or "account" in ll:
                user_labels.append(lbl)
            elif "growth" in ll:
                growth_labels.append(lbl)
            elif "mix" in ll or "% of" in ll:
                mix_labels.append(lbl)
            else:
                other_labels.append(lbl)

        if rev_labels:
            sub_header("Segment Revenue")
            for lbl in rev_labels:
                data_row(lbl, "SEG", lbl)

        if user_labels:
            sub_header("User Metrics / KPIs")
            for lbl in user_labels:
                data_row(lbl, "SEG", lbl)

        if growth_labels:
            sub_header("Segment Growth")
            for lbl in growth_labels:
                data_row(lbl, "SEG", lbl, FMT_PCT)

        if mix_labels:
            sub_header("Revenue Mix")
            for lbl in mix_labels:
                data_row(lbl, "SEG", lbl, FMT_PCT)

        if other_labels:
            sub_header("Other KPIs")
            for lbl in other_labels:
                # Detect format from label name
                fmt = FMT_NUMBER
                ll = lbl.lower()
                if "margin" in ll or "ratio" in ll:
                    fmt = FMT_PCT
                data_row(lbl, "SEG", lbl, fmt)

        blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # BALANCE SHEET (Internet)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("BALANCE SHEET")
    data_row("Cash & Equivalents",     "BS", "Cash")
    data_row("Total Assets",           "BS", "Total Assets")
    data_row("Goodwill & Intangibles", "BS", "Goodwill|Intangible")
    data_row("Total Debt",             "BS", "Total Debt")
    data_row("Net Cash / (Debt)",      "BS", "Net Cash")
    data_row("Shareholders' Equity",   "BS", "Equity")
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # VALUATION (Internet)
    # ════════════════════════════════════════════════════════════════════════════
    section_header("VALUATION")

    sub_header("Price Multiples")
    data_row("P/E",             "RAT", "P/E",   FMT_RATIO)
    data_row("P/FCF",           "RAT", "P/FCF", FMT_RATIO)
    data_row("P/S",             "RAT", "P/S",   FMT_RATIO)

    sub_header("EV Multiples")
    data_row("EV / Revenue",    "RAT", "EV/Revenue", FMT_RATIO)
    data_row("EV / EBITDA",     "RAT", "EV/EBITDA",  FMT_RATIO)
    data_row("EV / EBIT",       "RAT", "EV/EBIT",    FMT_RATIO)
    data_row("EV / FCF",        "RAT", "EV/FCF",     FMT_RATIO)

    sub_header("Returns")
    data_row("Return on Equity (ROE)",          "RAT", "Return on Equity",          FMT_PCT)
    data_row("Return on Assets (ROA)",          "RAT", "Return on Assets",          FMT_PCT)
    data_row("Return on Invested Capital (ROIC)","RAT","Return on Invested Capital", FMT_PCT)
    blank_row()

    # ════════════════════════════════════════════════════════════════════════════
    # BULL BEAR & TAILWINDS
    # ════════════════════════════════════════════════════════════════════════════
    _build_bull_bear_sheet(wb, bull_bear)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ─── Handler ──────────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            params = parse_qs(parsed.query)

            job_id_str = (params.get("jobId") or params.get("jobid") or [None])[0]
            if not job_id_str:
                self._json(400, {"error": "jobId query parameter is required"})
                return

            job_id = int(job_id_str)

            # 1. Fetch source sheets from DB
            sheets = fetch_sheets(job_id)
            if not sheets:
                self._json(404, {"error": f"No uploaded files found for jobId {job_id}"})
                return

            # 2. Detect company metadata
            company_type = detect_company_type(sheets)
            offsets, quarters, ttm_cols, ntm_cols = detect_offsets(sheets)
            ticker, company_name, company_id = fetch_company_info(job_id)

            # 3. Extract and store financial metrics in core_sheets for Claude
            extract_and_store_metrics(
                sheets, offsets, quarters, ttm_cols, company_id, company_type,
            )

            # 4. Fetch bull/bear thesis from DB (if previously generated)
            bull_bear = fetch_bull_bear(company_id)

            # 5. Build Excel — route to the correct template by company type
            build_fn_map = {
                "banking": build_banking_template,
                "financials": build_financials_template,
                "internet": build_internet_template,
            }
            build_fn = build_fn_map.get(company_type, build_software_template)
            excel_bytes = build_fn(
                sheets, offsets, quarters, ttm_cols, ntm_cols, ticker, company_name,
                bull_bear=bull_bear,
            )

            # 5. Stream response
            filename = f"{ticker}_CoreSheet.xlsx"
            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(excel_bytes)))
            self.end_headers()
            self.wfile.write(excel_bytes)

        except ValueError as e:
            self._json(400, {"error": str(e)})
        except Exception as e:
            self._json(500, {"error": str(e)})

    def _json(self, status: int, data: dict):
        body = json.dumps(data).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        pass  # suppress Vercel request logs
